"""
Training Crop Extractor (image crops + measured power labels)
=============================================================
Walks a "training data" folder. Each sub-folder is one LED array and contains:
  * a stitched array image (.png/.jpg/...)
  * an .xlsx workbook with an LIV sheet (name contains "LIV")

For every array it:
  1. Reads the LIV sheet and keeps only the rows measured at
         I_Low_mA == 0.75   and   VDD_50LED_Volt == 4.5
     taking the corresponding Power_uW for each Channel (e.g. "X02").
  2. Aligns the coordinate template to the image (same detect/align/refine/snap
     pipeline as the brightness meter) and crops each LED.
  3. Saves ONLY the LEDs that have a measured power, named by their alphabetical
     label + power, into a per-array output folder. LEDs without a measured
     power are skipped.

A channel like "X02" maps to label "X2" (letter = column, number = row position
in the staggered scheme - see led_label).  Output crop files are named
    <label>_<power>uW.png        e.g.  X2_27.302uW.png
and a manifest CSV ties every crop to its array / label / power.

Requirements:
    pip install opencv-python numpy openpyxl
"""

import cv2
import numpy as np
import csv
import os
import glob
import re
import openpyxl

# CONFIG  (paths are relative to this script's folder, so the project is portable)
BASE          = os.path.dirname(os.path.abspath(__file__))
TRAINING_DIR  = os.path.join(BASE, "training data")
TEMPLATE_CSV  = os.path.join(BASE, "array_coordinates_corrected.csv")
OUTPUT_FOLDER = os.path.join(BASE, "Training Crops")

# Which LIV measurement to label crops with:
I_LOW_TARGET  = 0.75   # mA
VDD_TARGET    = 4.5    # V (VDD_50LED_Volt)
TOL           = 1e-6   # float compare tolerance

# Crop size is RESOLUTION-INDEPENDENT: the square side is a fixed fraction of the
# detected LED pitch, so every LED is framed the same way regardless of whether the
# image is a high-res stitch or a smaller single shot.  CROP_SIZE is only a fallback
# (e.g. if the pitch can't be measured).
CROP_FRACTION = 0.60   # crop side = this fraction of the LED pitch (~115px at a 190px pitch)
CROP_SIZE     = 115    # fallback square side (px) if the LED pitch is unavailable
SAVE_COLOR    = True   # True -> colour crop; False -> blue channel only
PAD_EDGE      = True   # zero-pad edge crops so every crop is exactly crop_size^2

# Alignment is RIGID: the constant LED template is placed with a single
# similarity transform (rotation + uniform scale + translation) only. Boxes are
# never moved individually to chase the image - the whole grid moves together.
GATE_FRAC     = 0.40   # match radius as a fraction of LED spacing (< 0.5 so a
                       # grid point can only match its OWN LED, not a neighbour)

# The correct grid scale is fixed by the LED pitch (detected spacing / template
# spacing) - a robust, corner-independent estimate. We forbid the fit from
# shrinking the grid more than this fraction below that pitch-implied scale, so
# it can't collapse into a dense cluster and rack up bogus matches.
MAX_SCALE_DECREASE = 0.25   # e.g. 0.25 -> grid may be at most 25% smaller

# Step-2 grid offset: the angle/scale fit is reliable, but if a corner LED is
# off the grid can settle a row or column off. We align COLUMNS first, then ROWS,
# each by whole-cell slides, keeping the slide that lands the most grid points on
# real LEDs. A single-column slide carries the half-row stagger automatically
# (the template's column vector is ~half a row lower for each step), so odd
# column slides shift the grid ~100px vertically and even slides cancel it out.
MAX_COL_SHIFT  = 6     # search +/- this many COLUMNS
MAX_ROW_SHIFT  = 4     # search +/- this many ROWS
MIN_SHIFT_GAIN = 0.05  # only accept a shift if it lands >=5% more grid points on
                       # LEDs than no shift. A real cell-offset exposes a whole
                       # row/column (>10% gain); a spurious shift gains 1-2%, so
                       # this keeps an already-correct grid (and ambiguous sparse
                       # arrays) from being nudged onto a wrong-label position.

# Dot detection tuning
THRESH_FRAC   = 0.37
MIN_AREA      = 40


# --------------------------------------------------------------------------- #
#  Labelling                                                                   #
# --------------------------------------------------------------------------- #
def led_label(row, col):
    """Map (row, col) -> array label like 'C33'.

    Column letter:  cols 1-12 -> A-L, cols 13-24 -> N-Y  (M is skipped).
    Row number depends on the column group:
      "odd"  columns {1,3,5,7,9,11,14,16,18,20,22,24} use
             1,3,5,7,9,11,13,15,19,21,23,25,27,29,31,33  (odds 1-33, skipping 17)
      "even" columns (the rest) use
             2,4,6,8,10,12,14,16,18,20,22,24,26,28,30,32  (clean evens)
    Only 17 is ever skipped (in the odd group); the even group has no gap.
    """
    letters = "ABCDEFGHIJKLNOPQRSTUVWXY"          # 24 letters, M omitted
    odd_cols = {1, 3, 5, 7, 9, 11, 14, 16, 18, 20, 22, 24}
    odd_rows  = [1, 3, 5, 7, 9, 11, 13, 15, 19, 21, 23, 25, 27, 29, 31, 33]
    even_rows = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32]
    number = (odd_rows if col in odd_cols else even_rows)[row - 1]
    return f"{letters[col - 1]}{number}"


def pad_label(label):
    """Display form with a zero-padded 2-digit number: 'A1' -> 'A01', 'C33' -> 'C33'.

    Used only for drawing labels on overlay images; the CSV/filename labels keep
    their unpadded form."""
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", str(label))
    return f"{m.group(1)}{int(m.group(2)):02d}" if m else str(label)


def normalize_label(text):
    """'X02' or 'X2' -> ('X', 2). Returns None if it doesn't look like a label."""
    m = re.fullmatch(r"\s*([A-Za-z]+)\s*0*(\d+)\s*", str(text))
    if not m:
        return None
    return (m.group(1).upper(), int(m.group(2)))


def build_label_index(n_rows, n_cols):
    """Reverse map  (letter, number) -> (row, col)  for the whole grid."""
    idx = {}
    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            idx[normalize_label(led_label(r, c))] = (r, c)
    return idx


# --------------------------------------------------------------------------- #
#  Spreadsheet                                                                 #
# --------------------------------------------------------------------------- #
def read_liv_powers(xlsx_path):
    """Return {(letter, number): power_uW} for rows at I_LOW_TARGET / VDD_TARGET."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    liv = next((s for s in wb.sheetnames if "LIV" in s.upper()), None)
    if liv is None:
        wb.close()
        raise KeyError(f"no LIV sheet in {os.path.basename(xlsx_path)}")
    ws = wb[liv]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    col = {name: i for i, name in enumerate(header)}
    ci, ii = col.get("Channel", 0), col.get("I_Low_mA", 2)
    vi, pi = col.get("VDD_50LED_Volt", 3), col.get("Power_uW", 4)

    powers = {}
    for r in rows:
        ch = r[ci]
        if ch is None:
            continue
        try:
            il, vd, pw = float(r[ii]), float(r[vi]), float(r[pi])
        except (TypeError, ValueError):
            continue
        if abs(il - I_LOW_TARGET) < TOL and abs(vd - VDD_TARGET) < TOL:
            key = normalize_label(ch)
            if key is not None:
                powers[key] = pw
    wb.close()
    return powers, liv


# --------------------------------------------------------------------------- #
#  Alignment / cropping pipeline (shared with the brightness meter)            #
# --------------------------------------------------------------------------- #
def load_template(path):
    with open(path, newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)
    n_cols = len(header) // 2
    pts, labels = [], []
    for r_idx, row in enumerate(rows):
        for c_idx in range(n_cols):
            pts.append([float(row[c_idx * 2]), float(row[c_idx * 2 + 1])])
            labels.append((r_idx + 1, c_idx + 1))
    return np.array(pts, np.float32), n_cols, len(rows), labels


def detect_dots(blue):
    mx = float(blue.max())
    _, mask = cv2.threshold(blue.astype(np.uint8), int(mx * THRESH_FRAC), 255, cv2.THRESH_BINARY)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((5, 5), np.uint8))
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    centers = []
    for cnt in contours:
        if cv2.contourArea(cnt) < MIN_AREA:
            continue
        M = cv2.moments(cnt)
        if M["m00"]:
            centers.append([M["m10"] / M["m00"], M["m01"] / M["m00"]])
    return np.array(centers, np.float32)


def similarity_from_2pts(src, dst):
    """Rigid similarity (rotation + uniform scale + translation) mapping the
    src point pair onto the dst point pair."""
    (ax, ay), (bx, by) = src
    (Ax, Ay), (Bx, By) = dst
    vx, vy = bx - ax, by - ay
    Vx, Vy = Bx - Ax, By - Ay
    s = np.hypot(Vx, Vy) / np.hypot(vx, vy)
    ang = np.arctan2(Vy, Vx) - np.arctan2(vy, vx)
    c, sn = s * np.cos(ang), s * np.sin(ang)
    R = np.array([[c, -sn], [sn, c]])
    t = np.array([Ax, Ay]) - R @ np.array([ax, ay])
    M = np.zeros((2, 3), np.float32)
    M[:, :2] = R
    M[:, 2] = t
    return M


def _apply(M, pts):
    return cv2.transform(pts.reshape(-1, 1, 2), M).reshape(-1, 2)


def _nearest_dist(pts, detected):
    """Distance from each point in pts to the nearest detected dot."""
    d2 = ((pts[:, None, :] - detected[None, :, :]) ** 2).sum(axis=2)
    return np.sqrt(d2.min(axis=1))


def _median_spacing(points):
    """Median nearest-neighbour distance among points (the LED pitch)."""
    d2 = ((points[:, None, :] - points[None, :, :]) ** 2).sum(axis=2)
    np.fill_diagonal(d2, np.inf)
    return float(np.median(np.sqrt(d2.min(axis=1))))


def _scale_of(M):
    """Uniform scale factor of a similarity transform."""
    return float(np.hypot(M[0, 0], M[1, 0]))


def _corner_candidates(detected):
    """A handful of plausible top-left / top-right anchor dots.

    Robust to a missing corner LED: we try several extremes and let the global
    consensus score (in fit_rigid_transform) pick the right combination."""
    x, y = detected[:, 0], detected[:, 1]
    o_sum = np.argsort(x + y)        # top-left  -> small x+y
    o_diff = np.argsort(x - y)       # top-right -> large x-y
    o_x, o_y = np.argsort(x), np.argsort(y)
    tl = list(o_sum[:3]) + list(o_x[:2]) + list(o_y[:2])
    tr = list(o_diff[-3:]) + list(o_x[-2:]) + list(o_y[:2])
    uniq = lambda idx: [detected[i] for i in dict.fromkeys(int(j) for j in idx)]
    return uniq(tl), uniq(tr)


def _best_slide(base, deltas, detected, gate):
    """Given candidate whole-grid translations {k: delta}, pick the k whose grid
    lands the most points on LEDs. Only accept a non-zero k if it clears the
    gain threshold over k=0, else stay put. Returns (k, delta)."""
    inl = {k: int((_nearest_dist(base + d, detected) < gate).sum()) for k, d in deltas.items()}
    inl0 = inl[0]
    best = max(inl, key=lambda k: (inl[k], -abs(k)))   # max inliers, smaller |k| wins ties
    if inl[best] - inl0 < max(4, MIN_SHIFT_GAIN * inl0):
        best = 0
    return best, deltas[best]


def _gap_column_offset(M, template, n_cols, n_rows, detected):
    """Column offset from aligning the array's THREE black columns to the image.

    Every array has the same black structure horizontally:
      * the empty centre gap - the imaginary column 'N' between physical columns
        12 and 13, where the template has a double-width spacing, and
      * the two black edges just beyond the outermost LED columns (A and Y).
    No LEDs ever fall in the centre gap or outside the A..Y span, so these three
    empty strips are a far more reliable horizontal fiducial than counting LED
    overlaps. We map the detected LEDs into the template frame and pick the
    whole-column shift that puts the fewest LEDs in any black strip - i.e. the
    centre gap empty AND every LED inside the array edges. Returns +k (grid must
    move right k columns); ties prefer no shift."""
    tcolx = template.reshape(n_rows, n_cols, 2)[0, :, 0]
    gap_x = (tcolx[11] + tcolx[12]) / 2.0                 # imaginary column N (centre)
    colA, colY = tcolx[0], tcolx[-1]                      # outermost LED columns
    colw = float(np.median(np.diff(tcolx[:12])))          # column pitch (gap-free side)
    Minv = cv2.invertAffineTransform(M)
    xs = cv2.transform(np.asarray(detected, np.float32).reshape(-1, 1, 2),
                       Minv).reshape(-1, 2)[:, 0]         # LED x in template frame
    best_k, best_viol = 0, None
    for k in range(-MAX_COL_SHIFT, MAX_COL_SHIFT + 1):
        c, a, y = gap_x + k * colw, colA + k * colw, colY + k * colw
        viol = int(((xs > c - colw / 2) & (xs < c + colw / 2)).sum())   # LEDs in centre gap
        viol += int((xs < a - colw / 2).sum())                         # LEDs left of array
        viol += int((xs > y + colw / 2).sum())                         # LEDs right of array
        if best_viol is None or viol < best_viol or (viol == best_viol and abs(k) < abs(best_k)):
            best_viol, best_k = viol, k
    return best_k


def correct_grid_offset(M, template, n_cols, detected, gate):
    """Step 2: align COLUMNS (via the black 'column N' gap), then ROWS (by inliers).

    Columns are placed by aligning the always-empty gap column to the image's
    black column - robust even when the array's own LEDs are dim or ambiguous,
    where counting LED overlaps fails. The single-column slide uses the template's
    own column vector (carrying the ~half-row stagger). Rows are then slid one
    (un-staggered) row at a time, keeping the row with the most LED overlaps and a
    clear gain over no-shift.

    Returns (M, row_shift, col_shift); +row_shift = up rows, +col_shift = left columns."""
    base = _apply(M, template)
    base0 = base[0]                                            # row 1, col 1
    row_vec = _apply(M, template[n_cols:n_cols + 1])[0] - base0  # one row DOWN
    n_rows = len(template) // n_cols

    def col_disp(k):
        if k == 0:
            return np.zeros(2, np.float32)
        d = _apply(M, template[abs(k):abs(k) + 1])[0] - base0  # |k| columns RIGHT
        return d if k > 0 else -d                              # left = exact mirror

    # 1) columns: align the imaginary gap column ("N") to the image's black column
    kc = _gap_column_offset(M, template, n_cols, n_rows, detected)
    dcol = col_disp(kc)

    # 2) rows, from the column-corrected position (inlier search)
    row_deltas = {k: (k * row_vec).astype(np.float32) for k in range(-MAX_ROW_SHIFT, MAX_ROW_SHIFT + 1)}
    kr, drow = _best_slide(base + dcol, row_deltas, detected, gate)

    M2 = M.copy()
    M2[:, 2] += (dcol + drow).astype(M.dtype)
    return M2, -kr, -kc   # +row_shift = up rows, +col_shift = left columns


def fit_rigid_transform(template, n_cols, detected):
    """Best RIGID (similarity) transform placing the constant template on the
    detected LEDs.

    Step 1 - angle/scale/position: try several corner anchors and keep the one
    that lands the most template points on real LEDs (global consensus ->
    immune to the shifted-lattice trap), then one rigid polish.
    Step 2 - edge correction: snap the grid onto the array's true top/bottom and
    left/right borders.
    Returns (M, (tl, tr), (row_shift, col_shift))."""
    t_tl, t_tr = template[0], template[n_cols - 1]
    spacing = _median_spacing(detected)
    gate = GATE_FRAC * spacing
    # scale the grid SHOULD have, from the LED pitch; floor below which we won't
    # let it shrink (rejects the collapsed-grid degeneracy)
    ref_scale = spacing / _median_spacing(template)
    min_scale = (1.0 - MAX_SCALE_DECREASE) * ref_scale

    tl_cands, tr_cands = _corner_candidates(detected)
    best_M, best_inliers, best_corners = None, -1, None
    fallback_M, fallback_corners, fallback_gap = None, None, np.inf
    for a in tl_cands:
        for b in tr_cands:
            if np.allclose(a, b):
                continue
            M = similarity_from_2pts([t_tl, t_tr], [a, b])
            # keep a scale-sane fallback in case nothing clears the floor
            gap = abs(_scale_of(M) - ref_scale)
            if gap < fallback_gap:
                fallback_M, fallback_corners, fallback_gap = M, (a, b), gap
            if _scale_of(M) < min_scale:
                continue                      # too small -> reject collapse
            inliers = int((_nearest_dist(_apply(M, template), detected) < gate).sum())
            if inliers > best_inliers:
                best_M, best_inliers, best_corners = M, inliers, (a, b)

    if best_M is None:                        # every candidate was below the floor
        best_M, best_corners = fallback_M, fallback_corners

    # one rigid polish: re-fit a similarity to the consensus matches (keeps the
    # grid rigid - estimateAffinePartial2D has no shear/aspect freedom). The
    # scale floor still applies, so the polish can't collapse the grid either.
    M = best_M
    for _ in range(2):
        P = _apply(M, template)
        d = _nearest_dist(P, detected)
        keep = d < gate
        if keep.sum() < 12:
            break
        src = template[keep]
        dst = np.array([detected[np.argmin(((detected - p) ** 2).sum(1))]
                        for p in P[keep]], np.float32)
        M2, _ = cv2.estimateAffinePartial2D(src, dst, method=cv2.RANSAC,
                                            ransacReprojThreshold=6)
        if M2 is None or _scale_of(M2) < min_scale:
            break
        M = M2.astype(np.float32)

    # step 2: snap onto the true borders (fixes a row/column off from an unlit corner)
    M, row_shift, col_shift = correct_grid_offset(M, template, n_cols, detected, gate)
    return M, best_corners, (row_shift, col_shift)


def led_pitch(points):
    """Median nearest-neighbour distance among the aligned LED centres (LED pitch, px)."""
    pts = np.asarray(points, np.float64)
    if len(pts) < 2:
        return float(CROP_SIZE)
    nn = []
    for i in range(len(pts)):
        d = np.hypot(pts[:, 0] - pts[i, 0], pts[:, 1] - pts[i, 1])
        d[i] = np.inf
        nn.append(d.min())
    return float(np.median(nn))


def crop_size_for(points):
    """Resolution-independent square crop side (px): a fixed fraction of the LED pitch.

    Falls back to CROP_SIZE if the pitch can't be measured."""
    pitch = led_pitch(points)
    if not np.isfinite(pitch) or pitch <= 0:
        return CROP_SIZE
    return max(16, int(round(CROP_FRACTION * pitch)))


def crop_led(img, cx, cy, size, pad):
    h, w = img.shape[:2]
    half = size // 2
    x1, y1 = int(round(cx - half)), int(round(cy - half))
    x2, y2 = x1 + size, y1 + size
    if not pad:
        return img[max(0, y1):min(h, y2), max(0, x1):min(w, x2)]
    if img.ndim == 3:
        canvas = np.zeros((size, size, img.shape[2]), dtype=img.dtype)
    else:
        canvas = np.zeros((size, size), dtype=img.dtype)
    sx1, sy1 = max(0, x1), max(0, y1)
    sx2, sy2 = min(w, x2), min(h, y2)
    if sx2 <= sx1 or sy2 <= sy1:
        return canvas
    canvas[sy1 - y1:sy2 - y1, sx1 - x1:sx2 - x1] = img[sy1:sy2, sx1:sx2]
    return canvas


def align_image(img, template, n_cols):
    """Place the rigid template on the image. Returns (blue, aligned_pts, corners).

    aligned_pts are the per-LED centres from ONE rigid similarity transform;
    they are never nudged individually."""
    blue = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)[:, :, 2].astype(np.float32)
    detected = detect_dots(blue)
    if len(detected) < 4:
        return None, None, None, (0, 0)
    M, corners, shifts = fit_rigid_transform(template, n_cols, detected)
    aligned = _apply(M, template)
    return blue, aligned, corners, shifts


# --------------------------------------------------------------------------- #
#  Per-array processing                                                        #
# --------------------------------------------------------------------------- #
IMG_EXTS = ("*.png", "*.jpg", "*.jpeg", "*.bmp", "*.tif", "*.tiff")


def find_one(folder, patterns):
    for pat in patterns:
        hits = glob.glob(os.path.join(folder, pat))
        if hits:
            return sorted(hits)[0]
    return None


def process_array(array_dir, template, labels, n_cols, label_index, out_root):
    name = os.path.basename(array_dir.rstrip("\\/"))
    img_path = find_one(array_dir, IMG_EXTS)
    xlsx_path = find_one(array_dir, ("*.xlsx",))
    if img_path is None or xlsx_path is None:
        print(f"  ! {name}: missing image or xlsx - skipped")
        return []

    powers, liv_sheet = read_liv_powers(xlsx_path)
    if not powers:
        print(f"  ! {name}: no measurements at {I_LOW_TARGET}mA / {VDD_TARGET}V - skipped")
        return []

    img = cv2.imdecode(np.fromfile(img_path, dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        print(f"  ! {name}: could not read image - skipped")
        return []
    blue, aligned, corners, (row_shift, col_shift) = align_image(img, template, n_cols)
    if aligned is None:
        print(f"  ! {name}: too few dots detected - skipped")
        return []
    blue_u8 = np.clip(blue, 0, 255).astype(np.uint8)

    out_dir = os.path.join(out_root, name)
    os.makedirs(out_dir, exist_ok=True)

    crop_size = crop_size_for(aligned)     # pitch-relative, so crops are resolution-independent
    overlay = img.copy()
    half = crop_size // 2
    rows_written = []
    matched = set()
    for (row, col), (x, y) in zip(labels, aligned):
        key = normalize_label(led_label(row, col))
        if key not in powers:
            continue                       # this LED has no measured power -> skip
        matched.add(key)
        power = powers[key]
        label = led_label(row, col)

        source = img if SAVE_COLOR else blue_u8
        crop = crop_led(source, x, y, crop_size, PAD_EDGE)

        fname = f"{label}_{power}uW.png"
        cv2.imwrite(os.path.join(out_dir, fname), crop)
        rows_written.append({"array": name, "label": pad_label(label), "row": row, "col": col,
                             "x": int(round(x)), "y": int(round(y)),
                             "Power_uW": power, "crop_file": os.path.join(name, fname)})

        cv2.rectangle(overlay, (int(x) - half, int(y) - half),
                      (int(x) + half, int(y) + half), (0, 255, 0), 2)
        cv2.putText(overlay, pad_label(label), (int(x) - half, int(y) - half - 8),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 255), 3)

    for (x, y) in corners:
        cv2.circle(overlay, (int(x), int(y)), 34, (0, 0, 255), 4)

    # verification overlay (only the labelled LEDs are boxed)
    h, w = overlay.shape[:2]
    cv2.imwrite(os.path.join(out_root, f"{name}_verify.png"),
                cv2.resize(overlay, (min(1600, w), int(min(1600, w) * h / w))))

    # per-array manifest
    with open(os.path.join(out_dir, "labels.csv"), "w", newline="") as f:
        wri = csv.DictWriter(f, fieldnames=["array", "label", "row", "col",
                                            "x", "y", "Power_uW", "crop_file"])
        wri.writeheader(); wri.writerows(rows_written)

    missing = sorted(set(powers) - matched)
    extra = ""
    if missing:
        extra = f"  ({len(missing)} measured channels not found on grid: " \
                f"{', '.join(a + str(b) for a, b in missing)})"
    parts = []
    if row_shift > 0:   parts.append(f"up {row_shift} row(s)")
    elif row_shift < 0: parts.append(f"down {-row_shift} row(s)")
    if col_shift > 0:   parts.append(f"left {col_shift} col(s)")
    elif col_shift < 0: parts.append(f"right {-col_shift} col(s)")
    shift_note = f"  [shifted grid {', '.join(parts)}]" if parts else ""
    print(f"  {name}: {len(rows_written)}/{len(powers)} measured LEDs cropped "
          f"@ {crop_size}px (pitch {led_pitch(aligned):.0f}px){extra}{shift_note}")
    return rows_written


def main():
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    template, n_cols, n_rows, labels = load_template(TEMPLATE_CSV)
    label_index = build_label_index(n_rows, n_cols)
    print(f"Template: {n_cols} x {n_rows} = {len(template)} dots")
    print(f"Labelling crops at I_Low={I_LOW_TARGET}mA, VDD={VDD_TARGET}V  |  "
          f"crop {CROP_FRACTION:.0%} of LED pitch  |  "
          f"{'colour' if SAVE_COLOR else 'blue-channel'}\n")

    array_dirs = [d for d in sorted(glob.glob(os.path.join(TRAINING_DIR, "*")))
                  if os.path.isdir(d)]
    if not array_dirs:
        raise FileNotFoundError(f"No array sub-folders in {TRAINING_DIR}")
    print(f"Processing {len(array_dirs)} array(s):\n")

    all_rows = []
    for d in array_dirs:
        all_rows.extend(process_array(d, template, labels, n_cols, label_index, OUTPUT_FOLDER))

    manifest = os.path.join(OUTPUT_FOLDER, "training_manifest.csv")
    with open(manifest, "w", newline="") as f:
        wri = csv.DictWriter(f, fieldnames=["array", "label", "row", "col",
                                            "x", "y", "Power_uW", "crop_file"])
        wri.writeheader(); wri.writerows(all_rows)

    print(f"\nDone. {len(all_rows)} labelled crops across {len(array_dirs)} arrays.")
    print(f"Output: {OUTPUT_FOLDER}")
    print(f"Combined manifest: {manifest}")


if __name__ == "__main__":
    main()
